"""
Interfaz Tkinter para descargar las tasas pasivas de depósitos a plazo fijo
desde la SBS para un rango de fechas y exportarlas a Excel.
"""

from __future__ import annotations

import csv
import calendar
import ctypes
import os
import re
import sys
import time
import threading
import traceback
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Callable, Literal

import numpy as np
import pandas as pd
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.font as tkfont

# Workaround Python 3.13: el GC destruye tk.Variable desde hilos secundarios
_orig_var_del = tk.Variable.__del__
def _safe_var_del(self):
    try:
        _orig_var_del(self)
    except RuntimeError:
        pass
tk.Variable.__del__ = _safe_var_del

# ==================== KIT DE INTERFAZ SURA INVESTMENTS ====================
# Adaptado de la plantilla de interfaz gráfica compartida del equipo
# (componentes reutilizables: SuraEntry, SuraButton, SuraInputRow, SuraLogBox,
# SuraProgressBar, FontSet). Se mantiene aquí en el mismo archivo, siguiendo
# la convención de herramientas Python de una sola pieza usada en este
# proyecto (sin equipo de programación dedicado, para minimizar partes
# móviles al mantener con IA).

ButtonVariant = Literal["primary", "secondary"]
CommandFunction = Callable[[], None]


def resource_path(relative_path: str) -> str:
    candidates: list[str] = []

    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(meipass)

    here = os.path.abspath(os.path.dirname(__file__))
    candidates.append(here)                   # junto al script
    candidates.append(os.path.dirname(here))  # raiz del kit (assets/ y fonts/ compartidos)

    for base in candidates:
        full = os.path.join(base, relative_path)
        if os.path.exists(full):
            return full

    return os.path.join(here, relative_path)


def default_output_dir() -> Path:
    # Cuando se corre como .exe empaquetado (--onefile), os.getcwd() puede
    # apuntar a una carpeta inesperada según cómo se lance el ejecutable.
    # Se usa la carpeta del propio .exe (o del script) como valor por
    # defecto más predecible.
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def load_bundled_fonts() -> None:
    # Si "Sura Sans" no esta instalada en el equipo, carga las OTF de la carpeta
    # fonts/ como fuente privada del proceso (Windows), sin instalar nada.
    if sys.platform != "win32":
        return

    fonts_dir = resource_path("assets/fonts")
    if not os.path.isdir(fonts_dir):
        return

    try:
        FR_PRIVATE = 0x10
        gdi32 = ctypes.windll.gdi32
        gdi32.AddFontResourceExW.argtypes = [ctypes.c_wchar_p, ctypes.c_uint, ctypes.c_void_p]
        gdi32.AddFontResourceExW.restype = ctypes.c_int

        for name in os.listdir(fonts_dir):
            if name.lower().endswith((".otf", ".ttf")):
                try:
                    gdi32.AddFontResourceExW(os.path.join(fonts_dir, name), FR_PRIVATE, None)
                except Exception:
                    pass
    except Exception:
        pass


def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32":
        return

    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def is_windows_dark_mode() -> bool:
    if sys.platform != "win32":
        return False

    try:
        import winreg

        key_path = r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"

        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path) as key:
            value, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
            return value == 0

    except Exception:
        return False


# --- Identidad de esta herramienta ---
APP_NAME = "Descargador de Tasas Pasivas SBS"
APP_SUBTITLE = "FONDOS SURA SAF — depósitos a plazo fijo, portal SBS."

APP_WIDTH = 1040
APP_HEIGHT = 800

HEADER_HEIGHT = 150
ACCENT_HEIGHT = 2
MARGIN_X = 38

TITLE_Y = HEADER_HEIGHT + ACCENT_HEIGHT + 28
SEPARATOR_Y = TITLE_Y + 46
SUBTITLE_Y = SEPARATOR_Y + 16

FORM_Y = 292

PROGRESS_Y = 724
BUTTONS_Y = 744
LOG_BOTTOM_Y = PROGRESS_Y - 14  # el log siempre termina justo antes de la barra de progreso

COLOR_BLACK = "#161618"
COLOR_TEXT = "#24272A"
COLOR_TEXT_MUTED = "#59646E"
COLOR_BORDER = "#D9D9D9"
COLOR_BORDER_DARK = "#98A3AE"
COLOR_ROW_ALT = "#F6F6F6"
COLOR_WHITE = "#FFFFFF"
COLOR_BLUE = "#0A2CCE"
COLOR_DISABLED_BG = "#F4F4F4"
COLOR_DISABLED_TEXT = "#9A9A9A"

FONT_FALLBACK = "Arial"


def detect_sura_font(root: tk.Tk) -> str:
    try:
        families = list(tkfont.families(root))
        families_sorted = sorted(families, key=lambda name: name.lower())

        for family in families_sorted:
            normalized = family.lower().replace(" ", "")
            if "sura" in normalized:
                return family

    except Exception:
        pass

    return FONT_FALLBACK


class FontSet:
    def __init__(self, root: tk.Tk) -> None:
        self.brand: str = detect_sura_font(root)
        self.body: str = self.brand   # Sura Sans si esta disponible; Arial como fallback

        self.title: tkfont.Font = tkfont.Font(family=self.brand, size=21, weight="bold")
        self.subtitle: tkfont.Font = tkfont.Font(family=self.body, size=10)
        self.section: tkfont.Font = tkfont.Font(family=self.brand, size=12, weight="bold")
        self.label: tkfont.Font = tkfont.Font(family=self.body, size=10, weight="bold")
        self.input: tkfont.Font = tkfont.Font(family=self.body, size=10)
        self.button: tkfont.Font = tkfont.Font(family=self.brand, size=10, weight="bold")
        self.status: tkfont.Font = tkfont.Font(family=self.body, size=9)
        self.log: tkfont.Font = tkfont.Font(family=self.body, size=9)
        self.checkbox: tkfont.Font = tkfont.Font(family=self.body, size=9)


class SuraEntry(tk.Frame):
    def __init__(
        self,
        master: tk.Misc,
        width: int,
        height: int,
        font: tkfont.Font,
        readonly: bool = False,
    ) -> None:
        super().__init__(master, width=width, height=height, bg=COLOR_BORDER_DARK,
                          highlightthickness=0, bd=0)

        self.pack_propagate(False)
        self.grid_propagate(False)

        self.var: tk.StringVar = tk.StringVar()

        self.inner: tk.Frame = tk.Frame(self, bg=COLOR_WHITE, bd=0, highlightthickness=0)
        self.inner.pack(fill="both", expand=True, padx=1, pady=1)

        self.entry: tk.Entry = tk.Entry(
            self.inner, textvariable=self.var, bd=0, highlightthickness=0,
            relief="flat", bg=COLOR_WHITE, fg=COLOR_TEXT, insertbackground=COLOR_TEXT,
            font=font, justify="center",
        )
        self.entry.pack(fill="both", expand=True, padx=(10, 8), pady=(5, 4))

        self.entry.bind("<FocusIn>", self._on_focus_in)
        self.entry.bind("<FocusOut>", self._on_focus_out)

        if readonly:
            self.entry.configure(state="readonly", readonlybackground=COLOR_WHITE)

    def _on_focus_in(self, _event: tk.Event | None = None) -> None:
        self.configure(bg=COLOR_BLUE)

    def _on_focus_out(self, _event: tk.Event | None = None) -> None:
        self.configure(bg=COLOR_BORDER_DARK)

    def get(self) -> str:
        return self.var.get()

    def set(self, value: str) -> None:
        current_state = str(self.entry.cget("state"))
        if current_state == "readonly":
            self.entry.configure(state="normal")
            self.var.set(value)
            self.entry.configure(state="readonly")
        else:
            self.var.set(value)

    def clear(self) -> None:
        self.set("")

    def configure_validation(self, vcmd) -> None:
        self.entry.configure(validate="key", validatecommand=vcmd)


class SuraButton(tk.Frame):
    def __init__(
        self,
        master: tk.Misc,
        text: str,
        command: CommandFunction | None = None,
        variant: ButtonVariant = "secondary",
        width: int = 130,
        height: int = 36,
        font: tkfont.Font | None = None,
    ) -> None:
        self.variant: ButtonVariant = variant
        self.command: CommandFunction | None = command
        self.enabled: bool = True
        self.colors: dict[str, str] = self._resolve_colors(variant)

        super().__init__(master, width=width, height=height, bg=self.colors["border"],
                          bd=0, highlightthickness=0)

        self.pack_propagate(False)
        self.grid_propagate(False)

        self._text = text
        self.label: tk.Label = tk.Label(
            self, text=text, bg=self.colors["bg"], fg=self.colors["fg"],
            font=font, bd=0, cursor="hand2", anchor="center",
        )
        self.label.pack(fill="both", expand=True, padx=1, pady=1)

        self.label.bind("<Button-1>", self._on_click)
        self.label.bind("<Enter>", self._on_enter)
        self.label.bind("<Leave>", self._on_leave)

    @staticmethod
    def _resolve_colors(variant: ButtonVariant) -> dict[str, str]:
        if variant == "primary":
            return {
                "bg": COLOR_BLACK, "fg": COLOR_WHITE, "border": COLOR_BLACK,
                "hover": "#2A2A2D",
                "disabled_bg": COLOR_DISABLED_BG, "disabled_fg": COLOR_DISABLED_TEXT,
            }
        return {
            "bg": COLOR_WHITE, "fg": COLOR_TEXT, "border": COLOR_BLACK,
            "hover": "#F3F3F3",
            "disabled_bg": COLOR_DISABLED_BG, "disabled_fg": COLOR_DISABLED_TEXT,
        }

    def _on_click(self, _event: tk.Event | None = None) -> None:
        if not self.enabled:
            return
        if self.command is not None:
            self.command()

    def _on_enter(self, _event: tk.Event | None = None) -> None:
        if not self.enabled:
            return
        self.label.configure(bg=self.colors["hover"])

    def _on_leave(self, _event: tk.Event | None = None) -> None:
        if not self.enabled:
            return
        self.label.configure(bg=self.colors["bg"])

    def set_text(self, text: str) -> None:
        self._text = text
        self.label.configure(text=text)

    def set_enabled(self, enabled: bool) -> None:
        self.enabled = enabled
        if enabled:
            self.configure(bg=self.colors["border"])
            self.label.configure(bg=self.colors["bg"], fg=self.colors["fg"], cursor="hand2")
        else:
            self.configure(bg=COLOR_BORDER)
            self.label.configure(bg=self.colors["disabled_bg"], fg=self.colors["disabled_fg"],
                                  cursor="arrow")


class SuraInputRow(tk.Frame):
    def __init__(
        self,
        master: tk.Misc,
        label_text: str,
        fonts: FontSet,
        row_width: int,
        command: CommandFunction | None = None,
        button_text: str = "Seleccionar",
        row_bg: str = COLOR_WHITE,
    ) -> None:
        super().__init__(master, bg=row_bg, bd=0, highlightthickness=0)

        label_x: int = 24
        label_w: int = 210
        entry_x: int = 250
        button_w: int = 136
        right_pad: int = 24
        gap: int = 16

        control_h: int = 36
        control_y: int = (52 - control_h) // 2

        button_x: int = row_width - right_pad - button_w
        entry_w: int = button_x - gap - entry_x

        self.label: tk.Label = tk.Label(
            self, text=label_text, bg=row_bg, fg=COLOR_TEXT, font=fonts.label, anchor="w",
        )
        self.label.place(x=label_x, y=0, width=label_w, height=52)

        self.entry: SuraEntry = SuraEntry(self, width=entry_w, height=control_h, font=fonts.input)
        self.entry.place(x=entry_x, y=control_y)

        self.button: SuraButton = SuraButton(
            self, text=button_text, command=command, variant="secondary",
            width=button_w, height=control_h, font=fonts.button,
        )
        self.button.place(x=button_x, y=control_y)


class SuraLogBox(tk.Frame):
    def __init__(self, master: tk.Misc, fonts: FontSet, width: int, height: int) -> None:
        super().__init__(master, width=width, height=height, bg=COLOR_BORDER_DARK,
                          bd=0, highlightthickness=0)

        self.pack_propagate(False)
        self.grid_propagate(False)

        self.inner: tk.Frame = tk.Frame(self, bg=COLOR_WHITE, bd=0)
        self.inner.pack(fill="both", expand=True, padx=1, pady=1)

        self.text: tk.Text = tk.Text(
            self.inner, bd=0, highlightthickness=0, relief="flat",
            bg=COLOR_WHITE, fg=COLOR_TEXT, font=fonts.log, wrap="word",
        )
        self.scroll: tk.Scrollbar = tk.Scrollbar(self.inner, orient="vertical",
                                                  command=self.text.yview)
        self.text.configure(yscrollcommand=self.scroll.set)

        self.text.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=8)
        self.scroll.pack(side="right", fill="y")

    def write(self, message: str) -> None:
        self.text.insert("end", message.rstrip() + "\n")
        self.text.see("end")

    def clear(self) -> None:
        self.text.delete("1.0", "end")

    def set_content(self, lines: list[str]) -> None:
        self.clear()
        if lines:
            self.text.insert("end", "\n".join(lines) + "\n")
            self.text.see("end")

    def reposition(self, x: int, y: int, height: int) -> None:
        self.configure(height=height)
        self.place(x=x, y=y, height=height)


class SuraProgressBar(tk.Frame):
    def __init__(self, master: tk.Misc, width: int, height: int = 10) -> None:
        super().__init__(master, width=width, height=height, bg=COLOR_BORDER,
                          bd=0, highlightthickness=0)

        self.width_value: int = width
        self.height_value: int = height
        self.pack_propagate(False)

        self.fill: tk.Frame = tk.Frame(self, bg=COLOR_BLUE, bd=0, highlightthickness=0)
        self.fill.place(x=0, y=0, width=0, height=height)

    def set_progress(self, value: float) -> None:
        bounded_value: float = max(0.0, min(1.0, value))
        self.fill.place_configure(width=int(self.width_value * bounded_value))
# ==================== CONFIGURACIÓN BÁSICA ====================

URL = "https://www.sbs.gob.pe/app/pp/EstadisticasSAEEPortal/Paginas/TIPasivaDepositoEmpresa.aspx?tip=B"

INPUT_ID      = "ctl00_cphContent_rdpDate_dateInput"
BTN_ID        = "ctl00_cphContent_btnConsultar"
MAIN_TABLE_MN = "ctl00_cphContent_rpgActualPrimTablaMn_OT"
TAB_ME_ID     = "ctl00_cphContent_lbtnMex"
LBL_FECHA_ID  = "ctl00_cphContent_lblMensajeFecha"

WAIT_TIMEOUT = 20

RELEVANT_IDS_MN = {
    "ctl00_cphContent_rpgActualPrimTablaMn_OT",
    "ctl00_cphContent_rpgActualPrimTablaMn_ctl00_DataZone_DT",
    "ctl00_cphContent_rpgActualMn_OT",
    "ctl00_cphContent_rpgActualMn_ctl00_DataZone_DT",
}

RELEVANT_IDS_ME = {
    "ctl00_cphContent_rpgActualPrimTablaMex_OT",
    "ctl00_cphContent_rpgActualPrimTablaMex_ctl00_DataZone_DT",
    "ctl00_cphContent_rpgActualMex_OT",
    "ctl00_cphContent_rpgActualMex_ctl00_DataZone_DT",
}

RELEVANT_IDS = RELEVANT_IDS_MN | RELEVANT_IDS_ME

MAPPING_PARSE_MN = {
    "ctl00_cphContent_rpgActualPrimTablaMn_OT": "df_tabla10",
    "ctl00_cphContent_rpgActualMn_OT":          "df_tabla12",
}

MAPPING_PARSE_ME = {
    "ctl00_cphContent_rpgActualPrimTablaMex_OT": "df_tabla14",
    "ctl00_cphContent_rpgActualMex_OT":           "df_tabla16",
}

DESIRED_COLS = [
    "Banco",
    "Depósitos de Ahorro",
    "Hasta 30 días",
    "31-90 días",
    "91-180 días",
    "181-360 días",
    "Más de 360 días",
    "Depósitos a Plazo",
    "Depósitos CTS",
]

# ==================== MOTOR DE FERIADOS (Perú) ====================
#
# Se separa en tres capas:
#
# 1) FIXED_HOLIDAYS_MMDD: feriados nacionales de fecha fija que se repiten
#    todos los años por ley. Válido para cualquier año.
#
# 2) Feriados móviles (Jueves y Viernes Santo): se calculan a partir del
#    Domingo de Pascua (algoritmo de Meeus/Jones/Butcher), válido para
#    cualquier año en el calendario gregoriano.
#
# 3) ADDITIONAL_NONWORKING_DAYS: días no laborables que se decretan año a
#    año (normalmente vía Decreto Supremo) y NO siguen una regla calculable
#    (ej. bicentenarios, puentes administrativos). Esta tabla debe
#    actualizarse manualmente cada año cuando se publiquen los decretos
#    correspondientes; no hay forma de predecirla algorítmicamente.
#
# "es_nacional=True" => feriado general, afecta también al sector privado
#   (por lo tanto es esperable que la fuente SBS no tenga dato ese día).
# "es_nacional=False" => día no laborable declarado solo para el sector
#   público; el sector privado (y por ende la SBS) podría seguir operando
#   con normalidad, aunque en la práctica puede que igual no actualice.

FIXED_HOLIDAYS_MMDD: dict[tuple[int, int], str] = {
    (1, 1):   "Año Nuevo",
    (5, 1):   "Día del Trabajo",
    (6, 29):  "San Pedro y San Pablo",
    (7, 28):  "Fiestas Patrias (1er día)",
    (7, 29):  "Fiestas Patrias (2do día)",
    (8, 30):  "Santa Rosa de Lima",
    (10, 8):  "Combate de Angamos",
    (11, 1):  "Todos los Santos",
    (12, 8):  "Inmaculada Concepción",
    (12, 25): "Navidad",
}

# nombre, es_nacional
ADDITIONAL_NONWORKING_DAYS: dict[int, dict[tuple[int, int], tuple[str, bool]]] = {
    2025: {
        (6, 7):  ("Día de la Bandera", False),
        (7, 23): ("Día de las Fuerzas Armadas y PNP", False),
        (8, 6):  ("Bicentenario de la Batalla de Junín", True),
        (12, 9): ("Bicentenario de la Batalla de Ayacucho", True),
    },
    # Agregar aquí los decretos de años siguientes conforme se publiquen.
}


def easter_sunday(year: int) -> date:
    """Domingo de Pascua (calendario gregoriano). Algoritmo de Meeus/Jones/Butcher."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def get_holiday_info(d: date) -> tuple[bool, str | None, bool | None]:
    """
    Retorna (es_feriado, nombre, es_nacional) para la fecha dada.
    """
    mmdd = (d.month, d.day)

    if mmdd in FIXED_HOLIDAYS_MMDD:
        return True, FIXED_HOLIDAYS_MMDD[mmdd], True

    pascua = easter_sunday(d.year)
    jueves_santo = pascua - timedelta(days=3)
    viernes_santo = pascua - timedelta(days=2)
    if d == jueves_santo:
        return True, "Jueves Santo", True
    if d == viernes_santo:
        return True, "Viernes Santo", True

    extra = ADDITIONAL_NONWORKING_DAYS.get(d.year, {})
    if mmdd in extra:
        nombre, es_nacional = extra[mmdd]
        return True, nombre, es_nacional

    return False, None, None

# ==================== UTILIDADES DE FECHA ====================


def parse_ddmmyyyy(s: str) -> datetime:
    return datetime.strptime(s, "%d/%m/%Y")


def date_range(start_str: str, end_str: str):
    start = parse_ddmmyyyy(start_str)
    end   = parse_ddmmyyyy(end_str)
    if end < start:
        raise ValueError("end_date debe ser >= start_date")
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


BIMESTRE_MESES = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]


def bimestre_anterior(hoy: date) -> tuple[date, date]:
    """
    Retorna (inicio, fin) del bimestre par-impar (ene-feb, mar-abr, ...)
    anterior al que contiene la fecha "hoy". Ejemplo: ejecutado el 6 de
    julio o el 20 de agosto, ambos retornan mayo-junio del mismo año.
    """
    idx_actual = (hoy.month - 1) // 2
    idx_anterior = idx_actual - 1
    anio = hoy.year
    if idx_anterior < 0:
        idx_anterior = 5
        anio -= 1
    mes_inicio, mes_fin = BIMESTRE_MESES[idx_anterior]
    inicio = date(anio, mes_inicio, 1)
    ultimo_dia = calendar.monthrange(anio, mes_fin)[1]
    fin = date(anio, mes_fin, ultimo_dia)
    return inicio, fin

# ==================== PARSE DE TABLAS HTML ====================


def build_header_with_spans(table):
    thead = table.find("thead")
    if not thead:
        return None
    trs = thead.find_all("tr")
    total_cols = 0
    for cell in trs[0].find_all(["th", "td"]):
        total_cols += int(cell.get("colspan", 1))
    nrows = len(trs)
    grid  = [["" for _ in range(total_cols)] for _ in range(nrows)]

    def first_empty_col(row_idx):
        for c in range(total_cols):
            if grid[row_idx][c] == "":
                return c
        return None

    for r_idx, tr in enumerate(trs):
        for cell in tr.find_all(["th", "td"]):
            text    = cell.get_text(strip=True)
            colspan = int(cell.get("colspan", 1))
            rowspan = int(cell.get("rowspan", 1))
            c       = first_empty_col(r_idx)
            if c is None:
                continue
            for cc in range(colspan):
                for rr in range(rowspan):
                    if r_idx + rr < nrows and c + cc < total_cols:
                        if grid[r_idx + rr][c + cc] == "":
                            grid[r_idx + rr][c + cc] = text

    final_names = []
    for col in range(total_cols):
        parts   = [grid[r][col] for r in range(nrows) if grid[r][col]]
        compact = []
        prev    = None
        for p in parts:
            if p != prev:
                compact.append(p)
            prev = p
        final_names.append(" - ".join(compact) if compact else f"col_{col}")
    return final_names


def extract_inner_data(table):
    tbody = table.find("tbody")
    if tbody:
        rows = tbody.find_all("tr")
    else:
        thead = table.find("thead")
        skip  = len(thead.find_all("tr")) if thead else 0
        rows  = table.find_all("tr")[skip:]
    data = []
    for tr in rows:
        tds   = tr.find_all("td")
        cells = [td.get_text("\n", strip=True) for td in tds]
        if any(c.strip() != "" for c in cells):
            data.append(cells)
    return data


def extract_banks(main_table):
    bank_cells = main_table.find_all("td", class_="rpgRowHeaderField")
    banks = [td.get_text(strip=True) for td in bank_cells if td.get_text(strip=True) != ""]
    if not banks:
        for tr in main_table.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) == 1:
                text = tds[0].get_text(strip=True)
                if text:
                    banks.append(text)
    return banks


def clean_num(x):
    if x is None:
        return float("nan")
    s = str(x).strip()
    if s in ("", "-", "--", "NA", "N/A"):
        return float("nan")
    s2 = s.replace(",", "").replace(" ", "")
    try:
        return float(s2)
    except Exception:
        return s


def ensure_banco_as_column_local(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    df = df.copy()
    if "Banco" in df.columns:
        cols = list(df.columns)
        if cols[0] != "Banco":
            cols.remove("Banco")
            cols.insert(0, "Banco")
            df = df[cols]
        return df
    try:
        idx_vals = list(df.index.astype(str))
        non_num  = sum(1 for v in idx_vals if not re.match(r"^[\d\.\- ]+$", v))
        if non_num / max(1, len(idx_vals)) > 0.5:
            df = df.reset_index()
    except Exception:
        pass
    if "Banco" not in df.columns and len(df.columns) > 0:
        df = df.rename(columns={df.columns[0]: "Banco"})
    if "Banco" in df.columns:
        cols = list(df.columns)
        cols.remove("Banco")
        cols.insert(0, "Banco")
        df = df[cols]
    return df


def detect_mapping(df: pd.DataFrame | None) -> dict:
    if df is None:
        return {k: None for k in DESIRED_COLS}
    df       = ensure_banco_as_column_local(df.copy())
    existing = list(df.columns)

    def find_col_by_keywords(keywords):
        for col in existing:
            low = col.lower()
            for kw in keywords:
                if kw.lower() in low:
                    return col
        return None

    mapping                        = {}
    mapping["Banco"]               = "Banco" if "Banco" in existing else None
    mapping["Depósitos de Ahorro"] = find_col_by_keywords(
        ["ahorro", "depósitos de ahorro", "depositos de ahorro"]
    )
    mapping["Hasta 30 días"]       = find_col_by_keywords(["hasta 30", "hasta 30 días", "0-30"])
    mapping["31-90 días"]          = find_col_by_keywords(["31-90", "31 - 90"])
    mapping["91-180 días"]         = find_col_by_keywords(["91-180", "91 - 180"])
    mapping["181-360 días"]        = find_col_by_keywords(["181-360", "181 - 360"])
    mapping["Más de 360 días"]     = find_col_by_keywords(["más de 360", "mas de 360"])
    # Columna total "Depósitos a Plazo": coincidencia exacta para no capturar
    # las subcolumnas "Depósitos a Plazo - Hasta 30 días", etc.
    mapping["Depósitos a Plazo"]   = next(
        (c for c in existing if c.lower() in ("depósitos a plazo", "depositos a plazo")),
        None,
    )
    mapping["Depósitos CTS"]       = find_col_by_keywords(["cts", "depósitos cts", "depositos cts"])

    used      = set(v for v in mapping.values() if v)
    remaining = [c for c in existing if c not in used]
    for d in DESIRED_COLS:
        if d == "Banco":
            continue
        if not mapping.get(d):
            mapping[d] = remaining.pop(0) if remaining else None
    return mapping


def harmonize_to_desired(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    df       = ensure_banco_as_column_local(df.copy())
    existing = list(df.columns)
    if existing == DESIRED_COLS:
        return df[DESIRED_COLS].copy()
    mapping  = detect_mapping(df)
    n        = len(df)
    new_rows = {}
    for d in DESIRED_COLS:
        src         = mapping.get(d)
        new_rows[d] = df[src].values if (src and src in df.columns) else np.array([np.nan] * n)
    new_df = pd.DataFrame(new_rows)
    if "Banco" in new_df.columns:
        new_df["Banco"] = new_df["Banco"].astype(str).replace("nan", "")
    return new_df


def html_has_data(html_text: str, relevant_ids=None):
    soup   = BeautifulSoup(html_text, "html.parser")
    tables = soup.find_all("table")
    ids    = {t.get("id") for t in tables if t.get("id")}
    if relevant_ids is None:
        relevant_ids = RELEVANT_IDS
    return len(ids & relevant_ids) > 0, ids


def parse_tables_from_html(html_text: str, mapping: dict[str, str]):
    soup_local = BeautifulSoup(html_text, "html.parser")
    created:    dict[str, pd.DataFrame] = {}
    debug_info = []

    def align_header_and_rows(col_names, rows_clean):
        if not col_names:
            return col_names, rows_clean
        n_header = len(col_names)
        max_len  = max((len(r) for r in rows_clean), default=0)

        def looks_like_entity_header(h):
            h0 = (h or "").strip().lower()
            return (not h0) or ("empresa" in h0) or ("entidad" in h0) or ("sistema" in h0)

        if max_len == n_header - 1:
            if looks_like_entity_header(col_names[0]):
                col_names = col_names[1:]
                n_header -= 1
        elif max_len == n_header + 1:
            col_names  = col_names[1:]
            rows_clean = [r[1:] for r in rows_clean]
            n_header  -= 1

        fixed_rows = []
        for r in rows_clean:
            r2 = list(r[:n_header])
            if len(r2) < n_header:
                r2 += [""] * (n_header - len(r2))
            fixed_rows.append(r2)
        return col_names, fixed_rows

    for tabla_id, base_var in mapping.items():
        main_tbl = soup_local.find("table", {"id": tabla_id})
        if main_tbl is None:
            continue

        candidates = main_tbl.find_all("table")
        if candidates:
            inner_tbl = max(
                candidates,
                key=lambda t: sum(len(tr.find_all("td")) for tr in t.find_all("tr")),
            )
        else:
            inner_tbl = main_tbl

        col_names = build_header_with_spans(inner_tbl)
        rows      = extract_inner_data(inner_tbl)
        banks     = extract_banks(main_tbl)

        if not col_names:
            maxcols   = max((len(r) for r in rows), default=0)
            col_names = [f"col_{j}" for j in range(maxcols)]

        rows_clean = []
        for r in rows:
            r_clean = [cell.split("\n")[0].strip() if isinstance(cell, str) else cell for cell in r]
            rows_clean.append(r_clean)

        col_names, rows_clean = align_header_and_rows(col_names, rows_clean)

        if not rows_clean:
            df_inner = pd.DataFrame(columns=["Banco"] + col_names)
        else:
            df_inner = pd.DataFrame(rows_clean, columns=col_names)
            if banks and len(banks) == len(df_inner):
                df_inner.insert(0, "Banco", banks)
            else:
                df_inner.insert(0, "Banco", [""] * len(df_inner))

        for col in df_inner.columns:
            if col != "Banco":
                df_inner[col] = df_inner[col].apply(clean_num)

        df_inner = df_inner.reset_index(drop=True)
        created[base_var] = df_inner
        debug_info.append((base_var, df_inner.shape))

    return created, debug_info


def split_person_tables(df_persona: pd.DataFrame | None):
    if df_persona is None or not isinstance(df_persona, pd.DataFrame) or df_persona.empty:
        return None, None

    dfp      = ensure_banco_as_column_local(df_persona.copy())
    cols     = [c for c in dfp.columns if c != "Banco"]
    nat_cols = [c for c in cols if re.search(r"Natur", c, re.I)]
    jur_cols = [c for c in cols if re.search(r"Jur", c, re.I) or re.search(r"Jurid", c, re.I)]

    if not nat_cols and not jur_cols and cols:
        half     = len(cols) // 2
        nat_cols = cols[:half]
        jur_cols = cols[half:]

    def build_person_subdf(subcols):
        if not subcols:
            return None
        df2      = dfp[["Banco"] + subcols].copy()
        newnames = {}
        for c in subcols:
            if " - " in c:
                newnames[c] = c.split(" - ", 1)[1].strip()
            else:
                m = re.search(r"(Hasta.*|31-90.*|91-180.*|181-360.*|Más de.*)", c, re.I)
                newnames[c] = m.group(1).strip() if m else c
        df2 = df2.rename(columns=newnames)
        return ensure_banco_as_column_local(df2)

    return build_person_subdf(nat_cols), build_person_subdf(jur_cols)

# ==================== ARMADO DE FILAS PARA EXPORTACIÓN ====================


def build_general_rows_for_date(date_str: str, mn_general, me_general) -> pd.DataFrame | None:
    rows = []
    for df_src, moneda in [(mn_general, "Moneda Nacional"), (me_general, "Moneda Extranjera")]:
        if isinstance(df_src, pd.DataFrame):
            dh = harmonize_to_desired(df_src)
            if dh is not None and not dh.empty:
                df  = dh.copy()
                df.insert(0, "Tipo de Moneda", moneda)
                df.insert(0, "Fecha", date_str)
                cols = ["Fecha", "Tipo de Moneda", "Banco"] + [c for c in DESIRED_COLS if c != "Banco"]
                for c in cols:
                    if c not in df.columns:
                        df[c] = np.nan
                rows.append(df[cols])
    return pd.concat(rows, ignore_index=True) if rows else None


def build_person_rows_for_date(
    date_str: str, mn_nat, mn_jur, me_nat, me_jur
) -> pd.DataFrame | None:
    out_cols = [
        "Fecha", "Tipo de Moneda", "Tipo Persona", "Banco",
        "Hasta 30 días", "31-90 días", "91-180 días", "181-360 días", "Más de 360 días",
    ]
    rows = []

    def prepare_person_df(df, tipo_moneda, tipo_persona):
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            return None
        df2     = ensure_banco_as_column_local(df.copy())
        mapping = detect_mapping(df2)
        minimal = pd.DataFrame()
        minimal["Banco"] = df2["Banco"].astype(str)
        for c in ["Hasta 30 días", "31-90 días", "91-180 días", "181-360 días", "Más de 360 días"]:
            src         = mapping.get(c)
            minimal[c]  = df2[src].values if (src and src in df2.columns) else np.nan
        minimal.insert(0, "Tipo Persona",   tipo_persona)
        minimal.insert(0, "Tipo de Moneda", tipo_moneda)
        minimal.insert(0, "Fecha",          date_str)
        return minimal[out_cols]

    for dfp, moneda, persona in [
        (mn_nat, "Moneda Nacional",   "Natural"),
        (mn_jur, "Moneda Nacional",   "Jurídica"),
        (me_nat, "Moneda Extranjera", "Natural"),
        (me_jur, "Moneda Extranjera", "Jurídica"),
    ]:
        p = prepare_person_df(dfp, moneda, persona)
        if p is not None:
            rows.append(p)
    return pd.concat(rows, ignore_index=True) if rows else None

# ==================== UPSERT A EXCEL ====================


def upsert_to_excel_accum(path: str, new_df: pd.DataFrame, key_cols, sheet_name="Datos"):
    if new_df is None or new_df.empty:
        return "no_rows"

    if os.path.exists(path):
        try:
            existing = pd.read_excel(path, sheet_name=sheet_name)
        except Exception:
            try:
                existing = pd.read_excel(path)
            except Exception:
                existing = pd.DataFrame()
    else:
        existing = pd.DataFrame()

    def normalize_keys(df_keys: pd.DataFrame) -> pd.DataFrame:
        dfk = df_keys.copy()
        if "Fecha" in dfk.columns:
            dfk["Fecha"] = pd.to_datetime(dfk["Fecha"], dayfirst=True, errors="coerce").dt.date
        for c in dfk.columns:
            dfk[c] = dfk[c].astype(str)
        return dfk

    if existing.empty:
        merged = new_df.copy()
    else:
        if all(k in new_df.columns for k in key_cols) and all(k in existing.columns for k in key_cols):
            existing_keys     = normalize_keys(existing[key_cols])
            new_keys          = normalize_keys(new_df[key_cols])
            existing_key_tups = existing_keys.apply(tuple, axis=1)
            new_key_tups      = set(new_keys.apply(tuple, axis=1))
            keep_mask         = ~existing_key_tups.isin(new_key_tups)
            existing_kept     = existing[keep_mask].copy()
            merged            = pd.concat([existing_kept, new_df], ignore_index=True, sort=False)
        else:
            merged = pd.concat([existing, new_df], ignore_index=True, sort=False)

    if "Fecha" in merged.columns:
        merged["Fecha"] = pd.to_datetime(merged["Fecha"], dayfirst=True, errors="coerce")
        try:
            merged = merged.sort_values(["Fecha"]).reset_index(drop=True)
        except Exception:
            pass

    try:
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            merged.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets.get(sheet_name)
            if ws is None:
                return "written_no_ws"

            nrows = merged.shape[0] + 1
            ncols = merged.shape[1]

            if "Fecha" in merged.columns:
                try:
                    fecha_col_idx = merged.columns.get_loc("Fecha") + 1
                    for r in range(2, nrows + 1):
                        cell = ws.cell(row=r, column=fecha_col_idx)
                        if isinstance(cell.value, str):
                            try:
                                dt = pd.to_datetime(cell.value, dayfirst=True, errors="coerce")
                                if not pd.isna(dt):
                                    cell.value = dt.to_pydatetime()
                            except Exception:
                                pass
                        cell.number_format = "DD/MM/YYYY"
                except Exception:
                    pass

            for col_idx in range(1, ncols + 1):
                col_letter = get_column_letter(col_idx)
                header     = str(merged.columns[col_idx - 1])
                max_len    = len(header)
                try:
                    col_values  = merged.iloc[:, col_idx - 1].astype(str).fillna("")
                    max_val_len = col_values.map(len).max() if not col_values.empty else 0
                    max_len     = max(max_len, int(max_val_len))
                except Exception:
                    pass
                try:
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                except Exception:
                    pass

            try:
                last_col             = get_column_letter(ncols)
                ref                  = f"A1:{last_col}{nrows}"
                base_table_name      = (
                    "TasaPasivaTipoPersona" if "Tipo Persona" in merged.columns
                    else "TasaPasiva"
                )
                existing_table_names = set(ws.tables.keys())
                table_name_candidate = base_table_name
                i = 1
                while table_name_candidate in existing_table_names:
                    table_name_candidate = f"{base_table_name}_{i}"
                    i += 1
                table = Table(displayName=table_name_candidate, ref=ref)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False, showLastColumn=False,
                    showRowStripes=True,   showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            except Exception:
                pass

        return "written"
    except PermissionError as e:
        return f"error_write_permiso: {e}"
    except Exception as e:
        return f"error_write: {e}"

# ==================== HELPERS SELENIUM ====================

DATE_RE = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")


def get_effective_date_from_label(driver):
    try:
        txt = driver.find_element(By.ID, LBL_FECHA_ID).text.strip()
    except Exception:
        return None, None
    matches = DATE_RE.findall(txt)
    return (matches[-1] if matches else None), txt


def set_sbs_date(driver, date_str: str):
    """
    Establece la fecha via API JS de Telerik para actualizar el ClientState
    que el servidor lee en el postback.
    """
    wait     = WebDriverWait(driver, WAIT_TIMEOUT)
    attempts = 3
    for i in range(attempts):
        try:
            driver.execute_script(
                """
                var rdp = $find('ctl00_cphContent_rdpDate');
                if (rdp) {
                    var parts = arguments[0].split('/');
                    var d = new Date(parseInt(parts[2]), parseInt(parts[1]) - 1, parseInt(parts[0]));
                    rdp.set_selectedDate(d);
                }
                """,
                date_str,
            )
            time.sleep(0.3)
            inp = driver.find_element(By.ID, INPUT_ID)
            if (inp.get_attribute("value") or "").strip() != date_str:
                driver.execute_script(
                    """
                    var el = arguments[0];
                    var setter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value').set;
                    setter.call(el, arguments[1]);
                    el.dispatchEvent(new Event('input',  { bubbles: true }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                    el.blur();
                    """,
                    inp, date_str,
                )
                time.sleep(0.3)
            return
        except StaleElementReferenceException:
            if i == attempts - 1:
                raise
            time.sleep(0.5)
        except Exception:
            try:
                inp = wait.until(EC.element_to_be_clickable((By.ID, INPUT_ID)))
                driver.execute_script(
                    """
                    var el = arguments[0];
                    el.value = arguments[1];
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                    el.blur();
                    """,
                    inp, date_str,
                )
                time.sleep(0.4)
                return
            except Exception:
                if i == attempts - 1:
                    raise
                time.sleep(0.5)


def click_consultar_and_wait(driver, date_str: str):
    """
    Clic en Consultar y espera a que el label refleje la fecha solicitada
    en contexto de Moneda Nacional (postback MN completado).
    """
    wait     = WebDriverWait(driver, WAIT_TIMEOUT)
    attempts = 3
    old_eff, old_txt = get_effective_date_from_label(driver)

    for i in range(attempts):
        try:
            btn = wait.until(EC.element_to_be_clickable((By.ID, BTN_ID)))
            driver.execute_script("arguments[0].scrollIntoView(true);", btn)
            time.sleep(0.2)
            btn.click()

            def _mn_ready(drv):
                eff, txt = get_effective_date_from_label(drv)
                if eff is None:
                    return False
                return txt != old_txt or eff != old_eff

            try:
                wait.until(_mn_ready)
            except Exception:
                pass

            wait.until(EC.presence_of_element_located((By.ID, MAIN_TABLE_MN)))
            return get_effective_date_from_label(driver)

        except StaleElementReferenceException:
            if i == attempts - 1:
                raise
            time.sleep(0.5)
        except Exception:
            try:
                driver.execute_script(
                    "document.getElementById(arguments[0]).click();", BTN_ID
                )
                wait.until(EC.presence_of_element_located((By.ID, MAIN_TABLE_MN)))
                return get_effective_date_from_label(driver)
            except Exception:
                if i == attempts - 1:
                    raise
                time.sleep(0.5)

    return None, None


def click_me_tab_and_wait(driver, date_str: str) -> bool:
    """
    Clic en el tab Moneda Extranjera y espera a que el postback AJAX termine.

    Estrategia:
    1. Verificar que no hay postback en curso (isInAsyncPostBack = False).
    2. Verificar que el label dice "Extranjera... al date_str".
    Si el timeout expira sin confirmación, retorna False para que el caller
    no capture datos ME stale.
    """
    wait     = WebDriverWait(driver, WAIT_TIMEOUT)
    attempts = 3

    for i in range(attempts):
        try:
            mex_tab = wait.until(EC.element_to_be_clickable((By.ID, TAB_ME_ID)))
            mex_tab.click()

            def _me_ready(drv):
                # Primero verificar que el postback AJAX terminó
                try:
                    is_posting = drv.execute_script(
                        "try {"
                        "  var prm = Sys.WebForms.PageRequestManager.getInstance();"
                        "  return prm.get_isInAsyncPostBack();"
                        "} catch(e) { return false; }"
                    )
                    if is_posting:
                        return False
                except Exception:
                    pass
                # Luego verificar que el label refleja ME + fecha correcta
                eff, txt = get_effective_date_from_label(drv)
                if not txt or eff is None:
                    return False
                return "Extranjera" in txt and eff == date_str

            try:
                wait.until(_me_ready)
                time.sleep(0.3)
                return True
            except Exception:
                # El wait expiró: postback no confirmado, NO retornar True
                return False

        except StaleElementReferenceException:
            if i == attempts - 1:
                return False
            time.sleep(0.5)
        except Exception:
            if i == attempts - 1:
                return False
            time.sleep(0.5)
    return False


def consultar_y_parsear_fecha(
    driver, date_str: str, out_dir: str | None = None, log_fn=print
) -> dict:
    """
    Ejecuta la consulta completa contra la SBS para date_str (Paso 1:
    Consultar → MN, Paso 2: tab Extranjera → ME) y parsea las tablas
    resultantes. Levanta una excepción si la consulta en sí falla en
    cualquiera de sus pasos (fecha rechazada por la SBS, timeout, etc.).

    Retorna un dict con la misma forma que las entradas de la cache
    "ultimo_ok" de run_date_range, para poder reutilizar tanto el resultado
    de una fecha del rango como el de una fecha "semilla" anterior al rango.
    """
    set_sbs_date(driver, date_str)
    eff_date, lbl_txt = click_consultar_and_wait(driver, date_str)

    html_mn = driver.page_source

    ok_me = click_me_tab_and_wait(driver, date_str)
    if not ok_me:
        log_fn(f"[WARN] {date_str}: postback ME no confirmado. Datos ME omitidos para esta fecha.")
    html_me = driver.page_source if ok_me else None

    if out_dir is not None:
        date_suffix = datetime.strptime(date_str, "%d/%m/%Y").strftime("%d%m%Y")
        html_path = os.path.join(out_dir, f"sbs_fuente_{date_suffix}.html")
        html_for_file = html_me if html_me is not None else html_mn
        try:
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html_for_file)
        except Exception:
            pass

    mn_general = mn_nat = mn_jur = None
    tablas_mn_detectadas = 0
    has_mn, ids_mn = html_has_data(html_mn, RELEVANT_IDS_MN)
    if not has_mn:
        log_fn(f"[INFO] No hay tablas MN para {date_str}.")
    else:
        tablas_mn_detectadas = len(ids_mn & RELEVANT_IDS_MN)
        log_fn(f"[INFO] Tablas MN={tablas_mn_detectadas}  detectadas")
        mn_tables, created_mn = parse_tables_from_html(html_mn, MAPPING_PARSE_MN)
        for base_name, shape in created_mn:
            log_fn(f"  MN: {base_name} shape={shape}")
        mn_general = mn_tables.get("df_tabla10")
        mn_nat, mn_jur = split_person_tables(mn_tables.get("df_tabla12"))

    me_general = me_nat = me_jur = None
    tablas_me_detectadas = 0
    if html_me is None:
        log_fn(f"[INFO] {date_str}: ME omitido (postback no confirmado).")
    else:
        has_me, ids_me = html_has_data(html_me, RELEVANT_IDS_ME)
        if not has_me:
            log_fn(f"[INFO] No hay tablas ME para {date_str}.")
        else:
            tablas_me_detectadas = len(ids_me & RELEVANT_IDS_ME)
            log_fn(f"[INFO] Tablas ME={tablas_me_detectadas}  detectadas")
            me_tables, created_me = parse_tables_from_html(html_me, MAPPING_PARSE_ME)
            for base_name, shape in created_me:
                log_fn(f"  ME: {base_name} shape={shape}")
            me_general = me_tables.get("df_tabla14")
            me_nat, me_jur = split_person_tables(me_tables.get("df_tabla16"))

    return {
        "eff_date": eff_date,
        "mn_general": mn_general, "mn_nat": mn_nat, "mn_jur": mn_jur,
        "me_general": me_general, "me_nat": me_nat, "me_jur": me_jur,
        "tablas_mn": tablas_mn_detectadas, "tablas_me": tablas_me_detectadas,
    }




class DetailLogger:
    """
    Registra una fila por cada fecha procesada, pensado para poder filtrar
    después los casos donde no hubo dato pero debería haberlo habido
    (es_feriado_registrado = False y tuvo_datos = False), o donde se usó
    un valor arrastrado de un feriado no registrado en la base de datos.
    Usa ";" como separador para abrir directamente en Excel con configuración
    regional de Perú.
    """

    FIELDNAMES = [
        "Fecha", "DiaSemana", "EsFinDeSemana",
        "EsFeriadoRegistrado", "NombreFeriado", "EsNacional",
        "FechaEfectivaSBS", "HuboArrastre",
        "TablasMN", "TablasME",
        "FilasGeneral", "FilasPersona", "TuvoDatos",
        "Advertencias",
    ]

    def __init__(self, path: str):
        self.path = path
        self._header_escrito = os.path.exists(path)

    def log(self, **kwargs) -> None:
        fila = {k: kwargs.get(k, "") for k in self.FIELDNAMES}
        escribir_header = not self._header_escrito
        with open(self.path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=self.FIELDNAMES, delimiter=";")
            if escribir_header:
                writer.writeheader()
                self._header_escrito = True
            writer.writerow(fila)

# ==================== PROCESO PRINCIPAL ====================


def run_date_range(
    start_date_str:      str,
    end_date_str:        str,
    out_base_dir:        str | None = None,
    log_fn                          = print,
    simple_log_fn                   = None,
    import_natural:      bool = True,
    import_juridica:     bool = True,
    import_general:      bool = True,
    remove_html_after:   bool = False,
    skip_weekends:       bool = True,
    cancel_event: threading.Event | None = None,
    generate_detail_log: bool = False,
    progress_fn=None,
) -> str:
    """
    Retorna uno de: "completado", "cancelado", "error_chrome",
    "error_navegacion". Cualquier otro fallo se propaga como excepción.
    """
    if simple_log_fn is not None:
        simple_log_fn(f"Iniciando proceso desde {start_date_str} hasta {end_date_str}.")

    range_tag = f"{start_date_str.replace('/', '')}-{end_date_str.replace('/', '')}"
    base_dir  = out_base_dir or os.getcwd()
    out_dir   = os.path.join(base_dir, f"TPF_{range_tag}")
    os.makedirs(out_dir, exist_ok=True)

    file_general   = os.path.join(out_dir, f"TPF_General_{range_tag}.xlsx")
    file_by_person = os.path.join(out_dir, f"TPF_Persona_{range_tag}.xlsx")

    detail_logger = None
    if generate_detail_log:
        detail_path = os.path.join(out_dir, f"LOG_Detalle_{range_tag}.csv")
        detail_logger = DetailLogger(detail_path)
        log_fn(f"[INFO] Log detallado habilitado: {detail_path}")

    log_fn(f"[INFO] Carpeta de salida: {out_dir}")

    try:
        driver = webdriver.Chrome()
    except Exception as e:
        log_fn(f"[ERROR] No se pudo iniciar Google Chrome: {e}")
        log_fn("        Verificar que Chrome esté instalado y actualizado en el equipo.")
        if simple_log_fn is not None:
            simple_log_fn("No se pudo iniciar Chrome. Verificar instalación.")
        return "error_chrome"

    try:
        driver.get(URL)
        time.sleep(2)
    except Exception as e:
        log_fn(f"[ERROR] No se pudo cargar la página de la SBS: {e}")
        if simple_log_fn is not None:
            simple_log_fn("No se pudo cargar la página de la SBS. Verificar la conexión a internet.")
        try:
            driver.quit()
        except Exception:
            pass
        return "error_navegacion"

    overall_export_results = []
    fue_cancelado = False
    ultimo_ok: dict | None = None  # cache de la última fecha con datos reales (ver PASO 1)

    # Sembrar la cache con el último valor disponible ANTES del inicio del
    # rango. Sin esto, si el primer día del rango es justo un feriado real
    # (la SBS rechaza la consulta, no solo "no tiene dato nuevo"), no hay
    # nada de qué arrastrar y ese día queda sin fila. Se retrocede día a día
    # hasta encontrar una consulta que funcione (máximo 15 intentos).
    log_fn("[INFO] Buscando un valor previo al rango para poder completar el primer día si hiciera falta...")
    _semilla_dt = parse_ddmmyyyy(start_date_str) - timedelta(days=1)
    for _ in range(15):
        _semilla_str = _semilla_dt.strftime("%d/%m/%Y")
        try:
            _semilla = consultar_y_parsear_fecha(driver, _semilla_str, out_dir=None, log_fn=log_fn)
            if _semilla["mn_general"] is not None or _semilla["me_general"] is not None:
                ultimo_ok = {
                    "origen": _semilla["eff_date"] or _semilla_str,
                    "mn_general": _semilla["mn_general"], "mn_nat": _semilla["mn_nat"],
                    "mn_jur": _semilla["mn_jur"],
                    "me_general": _semilla["me_general"], "me_nat": _semilla["me_nat"],
                    "me_jur": _semilla["me_jur"],
                    "tablas_mn": _semilla["tablas_mn"], "tablas_me": _semilla["tablas_me"],
                }
                log_fn(f"[INFO] Valor semilla obtenido de {ultimo_ok['origen']}.")
                break
        except Exception:
            pass
        _semilla_dt -= timedelta(days=1)

    if ultimo_ok is None:
        log_fn(
            "[WARN] No se encontró ningún valor previo al rango para sembrar el arrastre. "
            "Si el primer día del rango no tiene dato propio, quedará sin fila."
        )

    _total_dias = max(
        1, (parse_ddmmyyyy(end_date_str).date() - parse_ddmmyyyy(start_date_str).date()).days + 1
    )
    _dias_procesados = 0

    try:
        for dt in date_range(start_date_str, end_date_str):

            _dias_procesados += 1
            if progress_fn is not None:
                try:
                    progress_fn(_dias_procesados / _total_dias)
                except Exception:
                    pass

            if cancel_event is not None and cancel_event.is_set():
                fue_cancelado = True
                log_fn("")
                log_fn("[INFO] Proceso detenido por el usuario. Cancelando el resto del rango.")
                if simple_log_fn is not None:
                    simple_log_fn("Proceso detenido por el usuario.")
                break

            date_str   = dt.strftime("%d/%m/%Y")
            is_weekend = dt.weekday() >= 5
            advertencias: list[str] = []

            if skip_weekends and is_weekend:
                log_fn(f"[INFO] Omitiendo fin de semana {date_str}.")
                if simple_log_fn is not None:
                    simple_log_fn(f"{date_str}: omitido (fin de semana).")
                if detail_logger:
                    detail_logger.log(
                        Fecha=date_str, DiaSemana=dt.strftime("%A"),
                        EsFinDeSemana=True, TuvoDatos=False,
                    )
                continue

            es_feriado, nombre_feriado, es_nacional = get_holiday_info(dt.date())
            if es_feriado:
                tipo = "feriado nacional" if es_nacional else "día no laborable (sector público)"
                log_fn(
                    f"[INFO] {date_str}: {nombre_feriado} — {tipo} (registrado en la base de "
                    f"datos). Se consulta igual para completar con el último valor disponible."
                )

            date_suffix       = dt.strftime("%d%m%Y")
            had_data_for_date = False

            log_fn("")
            log_fn("=" * 60)
            log_fn(f"Procesando fecha {date_str} (sufijo: {date_suffix})")
            log_fn("=" * 60)

            uso_cache = False
            origen_arrastre = date_str

            # ===== Consultar la fecha (Paso 1 MN + Paso 2 ME + parseo) =====
            try:
                resultado = consultar_y_parsear_fecha(driver, date_str, out_dir=out_dir, log_fn=log_fn)
                eff_date   = resultado["eff_date"]
                mn_general = resultado["mn_general"]
                mn_nat     = resultado["mn_nat"]
                mn_jur     = resultado["mn_jur"]
                me_general = resultado["me_general"]
                me_nat     = resultado["me_nat"]
                me_jur     = resultado["me_jur"]
                tablas_mn_detectadas = resultado["tablas_mn"]
                tablas_me_detectadas = resultado["tablas_me"]
            except Exception as e:
                # En un feriado real la SBS puede directamente rechazar la
                # consulta (el datepicker no habilita "Consultar", o el
                # postback nunca completa) en vez de responder con la fecha
                # vigente anterior. Lo mismo puede pasar por cualquier otro
                # fallo puntual de red/render. En ese caso no sirve esperar a
                # que la SBS "avise" la fecha vigente (mecanismo usado más
                # abajo): hay que arrastrar directamente desde la última
                # consulta exitosa que se tiene en memoria (incluida la
                # "semilla" previa al rango), sin volver a tocar el navegador.
                if ultimo_ok is not None:
                    uso_cache = True
                    origen_arrastre = ultimo_ok["origen"]
                else:
                    log_fn(f"[ERROR] No se pudo consultar la fecha {date_str}: {e}")
                    if simple_log_fn is not None:
                        simple_log_fn(f"{date_str}: error al consultar en la SBS.")
                    if detail_logger:
                        detail_logger.log(
                            Fecha=date_str, DiaSemana=dt.strftime("%A"),
                            EsFinDeSemana=False, EsFeriadoRegistrado=es_feriado,
                            NombreFeriado=nombre_feriado or "",
                            EsNacional=es_nacional if es_nacional is not None else "",
                            TuvoDatos=False,
                            Advertencias=f"Error al consultar y sin dato previo para arrastrar: {e}",
                        )
                    continue

            hubo_arrastre = False

            if uso_cache:
                hubo_arrastre = True
                mn_general = ultimo_ok["mn_general"]
                mn_nat     = ultimo_ok["mn_nat"]
                mn_jur     = ultimo_ok["mn_jur"]
                me_general = ultimo_ok["me_general"]
                me_nat     = ultimo_ok["me_nat"]
                me_jur     = ultimo_ok["me_jur"]
                tablas_mn_detectadas = ultimo_ok["tablas_mn"]
                tablas_me_detectadas = ultimo_ok["tablas_me"]

                if es_feriado:
                    msg = (
                        f"{date_str}: {nombre_feriado} ({tipo}). La SBS no permitió consultar "
                        f"esta fecha; se completa con el valor de {origen_arrastre}."
                    )
                    log_fn(f"[INFO] {msg}")
                    if simple_log_fn is not None:
                        simple_log_fn(f"{date_str}: {nombre_feriado} — se usó el valor de {origen_arrastre}.")
                else:
                    msg = (
                        f"{date_str}: no se pudo consultar en la SBS y esta fecha NO está "
                        f"registrada como feriado/no laborable en la base de datos interna. "
                        f"Se completa con el valor de {origen_arrastre}. Evaluar si corresponde "
                        f"agregar esta fecha a la base de feriados."
                    )
                    log_fn(f"[WARN] {msg}")
                    advertencias.append(msg)
                    if simple_log_fn is not None:
                        simple_log_fn(
                            f"{date_str}: ⚠ error al consultar y feriado no registrado — "
                            f"se usó el valor de {origen_arrastre}."
                        )

            elif eff_date and eff_date != date_str:
                # Sin dato propio para esta fecha, pero la SBS sí respondió con
                # la fecha vigente anterior (a diferencia del caso de arriba,
                # donde directamente rechaza la consulta). Se arrastra ese valor.
                hubo_arrastre = True
                origen_arrastre = eff_date
                if es_feriado:
                    msg = (
                        f"{date_str}: {nombre_feriado} ({tipo}). Sin dato propio en la SBS; "
                        f"se completa con el valor de {eff_date}."
                    )
                    log_fn(f"[INFO] {msg}")
                    if simple_log_fn is not None:
                        simple_log_fn(f"{date_str}: {nombre_feriado} — se usó el valor de {eff_date}.")
                else:
                    msg = (
                        f"{date_str}: la SBS no tiene data propia; indica vigente al {eff_date}. "
                        f"Esta fecha NO está registrada como feriado/no laborable en la base de "
                        f"datos interna. Se usará el valor de {eff_date} para completar {date_str}. "
                        f"Evaluar si corresponde agregar esta fecha a la base de feriados."
                    )
                    log_fn(f"[WARN] {msg}")
                    advertencias.append(msg)
                    if simple_log_fn is not None:
                        simple_log_fn(
                            f"{date_str}: ⚠ feriado no registrado en la base de datos — "
                            f"se usó el valor de {eff_date}."
                        )

            # ===== Armar y exportar =====
            general_rows = (
                build_general_rows_for_date(date_str, mn_general, me_general)
                if import_general else None
            )
            person_rows = build_person_rows_for_date(date_str, mn_nat, mn_jur, me_nat, me_jur)

            if person_rows is not None:
                if not import_natural:
                    person_rows = person_rows[
                        person_rows["Tipo Persona"] != "Natural"
                    ].reset_index(drop=True)
                if not import_juridica:
                    person_rows = person_rows[
                        person_rows["Tipo Persona"] != "Jurídica"
                    ].reset_index(drop=True)
                if person_rows.empty:
                    person_rows = None

            filas_general_n = 0
            filas_persona_n = 0

            if general_rows is not None and not general_rows.empty:
                had_data_for_date = True
                filas_general_n = len(general_rows)
                res = upsert_to_excel_accum(
                    file_general, general_rows, ["Fecha", "Tipo de Moneda", "Banco"]
                )
                overall_export_results.append(
                    (file_general, date_suffix, res, f"rows={filas_general_n}")
                )
                if isinstance(res, str) and res.startswith("error_write"):
                    msg = (
                        f"No se pudo escribir {os.path.basename(file_general)}: {res}. "
                        f"Verificar que el archivo no esté abierto en Excel u otra aplicación."
                    )
                    log_fn(f"[ERROR] {msg}")
                    advertencias.append(msg)
                    if simple_log_fn is not None:
                        simple_log_fn(
                            f"{date_str}: ⚠ error al guardar en "
                            f"{os.path.basename(file_general)} (¿archivo abierto?)."
                        )
                else:
                    log_fn(
                        f"[EXPORT] General -> {os.path.basename(file_general)}: "
                        f"{res} ({filas_general_n} filas)"
                    )

            if person_rows is not None and not person_rows.empty:
                had_data_for_date = True
                filas_persona_n = len(person_rows)
                res2 = upsert_to_excel_accum(
                    file_by_person, person_rows,
                    ["Fecha", "Tipo de Moneda", "Tipo Persona", "Banco"],
                )
                overall_export_results.append(
                    (file_by_person, date_suffix, res2, f"rows={filas_persona_n}")
                )
                if isinstance(res2, str) and res2.startswith("error_write"):
                    msg = (
                        f"No se pudo escribir {os.path.basename(file_by_person)}: {res2}. "
                        f"Verificar que el archivo no esté abierto en Excel u otra aplicación."
                    )
                    log_fn(f"[ERROR] {msg}")
                    advertencias.append(msg)
                    if simple_log_fn is not None:
                        simple_log_fn(
                            f"{date_str}: ⚠ error al guardar en "
                            f"{os.path.basename(file_by_person)} (¿archivo abierto?)."
                        )
                else:
                    log_fn(
                        f"[EXPORT] Por tipo de persona -> {os.path.basename(file_by_person)}: "
                        f"{res2} ({filas_persona_n} filas)"
                    )

            # Si esta fecha se resolvió con datos frescos (no desde cache), la
            # guardamos como la nueva "última consulta exitosa" para que fechas
            # futuras que fallen (feriados reales, errores puntuales) puedan
            # arrastrar desde acá sin depender de que la SBS responda.
            if not uso_cache and had_data_for_date:
                ultimo_ok = {
                    "origen": origen_arrastre,
                    "mn_general": mn_general, "mn_nat": mn_nat, "mn_jur": mn_jur,
                    "me_general": me_general, "me_nat": me_nat, "me_jur": me_jur,
                    "tablas_mn": tablas_mn_detectadas, "tablas_me": tablas_me_detectadas,
                }

            if simple_log_fn is not None:
                if had_data_for_date:
                    simple_log_fn(f"{date_str} descargado exitosamente ✅")
                else:
                    simple_log_fn(f"{date_str}: sin datos disponibles.")

            if detail_logger:
                detail_logger.log(
                    Fecha=date_str, DiaSemana=dt.strftime("%A"),
                    EsFinDeSemana=False, EsFeriadoRegistrado=es_feriado,
                    NombreFeriado=nombre_feriado or "",
                    EsNacional=es_nacional if es_nacional is not None else "",
                    FechaEfectivaSBS=origen_arrastre,
                    HuboArrastre=hubo_arrastre,
                    TablasMN=tablas_mn_detectadas, TablasME=tablas_me_detectadas,
                    FilasGeneral=filas_general_n, FilasPersona=filas_persona_n,
                    TuvoDatos=had_data_for_date,
                    Advertencias="; ".join(advertencias),
                )

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    log_fn("")
    log_fn("Resumen exportaciones:")
    for fp, suf, action, details in overall_export_results:
        log_fn(f" - {os.path.basename(fp)} | sufijo={suf} | {action} | {details}")
    if not overall_export_results:
        log_fn(" - No se encontraron datos en el rango especificado.")

    if remove_html_after:
        removed = 0
        for fname in os.listdir(out_dir):
            if fname.startswith("sbs_fuente_") and fname.endswith(".html"):
                try:
                    os.remove(os.path.join(out_dir, fname))
                    removed += 1
                except Exception:
                    pass
        log_fn(f"[INFO] Eliminados {removed} archivos temporales HTML.")

    if progress_fn is not None and not fue_cancelado:
        try:
            progress_fn(1.0)
        except Exception:
            pass

    return "cancelado" if fue_cancelado else "completado"
# ==================== INTERFAZ TKINTER ====================


class TasasApp:
    def __init__(self) -> None:
        enable_windows_dpi_awareness()
        load_bundled_fonts()

        self.root: tk.Tk = tk.Tk()
        self.root.title(f"SURA INVESTMENTS | {APP_NAME}")
        self.root.geometry(f"{APP_WIDTH}x{APP_HEIGHT}")
        self.root.resizable(False, False)
        self.root.minsize(APP_WIDTH, APP_HEIGHT)
        self.root.maxsize(APP_WIDTH, APP_HEIGHT)
        self.root.configure(bg=COLOR_WHITE)

        self.fonts: FontSet = FontSet(self.root)
        self._images: list[tk.PhotoImage] = []

        self.worker: threading.Thread | None = None
        self.cancel_event = threading.Event()
        self._worker_status: str | None = None

        self._log_simple_lines: list[str] = []
        self._log_advanced_lines: list[str] = []
        self._console_mode: str = "simple"

        self._set_window_icon()
        self._build_ui()
        self._center_window()
        self._set_default_dates()

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # ---------------------------------------------------------------- infra

    def _center_window(self) -> None:
        screen_width: int = self.root.winfo_screenwidth()
        screen_height: int = self.root.winfo_screenheight()
        pos_x: int = max(0, (screen_width - APP_WIDTH) // 2)
        pos_y: int = max(0, (screen_height - APP_HEIGHT) // 2)
        self.root.geometry(f"{APP_WIDTH}x{APP_HEIGHT}+{pos_x}+{pos_y}")

    def _load_photo(self, relative_path: str) -> tk.PhotoImage | None:
        path: str = resource_path(relative_path)
        if not os.path.exists(path):
            return None
        try:
            image: tk.PhotoImage = tk.PhotoImage(file=path)
            self._images.append(image)
            return image
        except Exception:
            return None

    def _set_window_icon(self) -> None:
        # Ala según el tema de la PC donde se ejecuta:
        #   modo claro  -> ala negra  (app_dark.ico)
        #   modo oscuro -> ala blanca (app_light.ico)
        dark_mode: bool = is_windows_dark_mode()
        ico_name: str = "assets/icons/app_light.ico" if dark_mode else "assets/icons/app_dark.ico"
        ico_path: str = resource_path(ico_name)

        if sys.platform == "win32":
            if os.path.exists(ico_path):
                try:
                    self.root.iconbitmap(default=ico_path)
                except Exception:
                    pass
            return

        png_name: str = "assets/icons/favicon_light.png" if dark_mode else "assets/icons/favicon_dark.png"
        icon: tk.PhotoImage | None = self._load_photo(png_name)
        if icon is None:
            icon = self._load_photo("assets/icons/favicon.png")
        if icon is not None:
            try:
                self.root.iconphoto(True, icon)
            except Exception:
                pass

    # ---------------------------------------------------------------- build

    def _build_ui(self) -> None:
        self._build_header()
        self._build_title_area()
        self._build_form_panel()
        self._build_advanced_toggle()
        self._build_log_area()
        self._build_action_bar()
        self._reposition_log_area()

    def _build_header(self) -> None:
        header: tk.Canvas = tk.Canvas(
            self.root, width=APP_WIDTH, height=HEADER_HEIGHT,
            bg=COLOR_BLACK, bd=0, highlightthickness=0,
        )
        header.place(x=0, y=0)

        banner: tk.PhotoImage | None = self._load_photo("assets/banners/header.png")
        if banner is not None:
            header.create_image(0, 0, image=banner, anchor="nw")
        else:
            header.create_text(
                APP_WIDTH // 2, HEADER_HEIGHT // 2, text="SURA  INVESTMENTS",
                fill=COLOR_WHITE, font=(self.fonts.brand, 26, "bold"),
            )

        accent: tk.Frame = tk.Frame(self.root, bg=COLOR_BLUE)
        accent.place(x=0, y=HEADER_HEIGHT, width=APP_WIDTH, height=ACCENT_HEIGHT)

    def _build_title_area(self) -> None:
        tk.Label(
            self.root, text=APP_NAME, bg=COLOR_WHITE, fg=COLOR_BLACK,
            font=self.fonts.title, anchor="w",
        ).place(x=MARGIN_X, y=TITLE_Y, width=APP_WIDTH - 2 * MARGIN_X, height=36)

        tk.Frame(self.root, bg=COLOR_BORDER).place(
            x=MARGIN_X, y=SEPARATOR_Y, width=APP_WIDTH - 2 * MARGIN_X, height=1,
        )

        tk.Label(
            self.root, text=APP_SUBTITLE, bg=COLOR_WHITE, fg=COLOR_TEXT,
            font=self.fonts.subtitle, anchor="w",
        ).place(x=MARGIN_X, y=SUBTITLE_Y, width=APP_WIDTH - 2 * MARGIN_X, height=24)

    def _build_form_panel(self) -> None:
        panel_width: int = APP_WIDTH - 2 * MARGIN_X
        inner_width: int = panel_width - 2

        # Alturas de fila: fechas (64) + carpeta de salida (52) + importación (52)
        row_h_dates = 64
        row_h_folder = 52
        row_h_import = 52
        self._form_height = row_h_dates + row_h_folder + row_h_import

        form_panel: tk.Frame = tk.Frame(self.root, bg=COLOR_BORDER, bd=0, highlightthickness=0)
        form_panel.place(x=MARGIN_X, y=FORM_Y, width=panel_width, height=self._form_height + 2)

        inner: tk.Frame = tk.Frame(form_panel, bg=COLOR_WHITE, bd=0, highlightthickness=0)
        inner.place(x=1, y=1, width=inner_width, height=self._form_height)

        self._build_dates_row(inner, inner_width, row_h_dates, y=0, row_bg=COLOR_WHITE)

        folder_row = SuraInputRow(
            inner, label_text="Carpeta de salida", fonts=self.fonts, row_width=inner_width,
            command=self._choose_folder, button_text="Cambiar...", row_bg=COLOR_ROW_ALT,
        )
        folder_row.place(x=0, y=row_h_dates, width=inner_width, height=row_h_folder)
        self.folder_entry = folder_row.entry
        self.folder_entry.set(str(default_output_dir()))

        self._build_import_row(
            inner, inner_width, row_h_import, y=row_h_dates + row_h_folder, row_bg=COLOR_WHITE,
        )

        self._form_end_y = FORM_Y + self._form_height + 2

    def _build_dates_row(self, parent: tk.Misc, row_width: int, row_h: int, y: int, row_bg: str) -> None:
        row = tk.Frame(parent, bg=row_bg, bd=0, highlightthickness=0)
        row.place(x=0, y=y, width=row_width, height=row_h)

        label_x = 24
        control_h = 36
        control_y = (row_h - control_h) // 2

        tk.Label(
            row, text="Fecha inicio", bg=row_bg, fg=COLOR_TEXT, font=self.fonts.label, anchor="w",
        ).place(x=label_x, y=0, width=110, height=row_h)

        x = label_x + 110
        self.s_day = SuraEntry(row, width=44, height=control_h, font=self.fonts.input)
        self.s_day.place(x=x, y=control_y); x += 44 + 4
        tk.Label(row, text="/", bg=row_bg, fg=COLOR_TEXT_MUTED, font=self.fonts.input).place(
            x=x, y=control_y, width=10, height=control_h); x += 14
        self.s_month = SuraEntry(row, width=44, height=control_h, font=self.fonts.input)
        self.s_month.place(x=x, y=control_y); x += 44 + 4
        tk.Label(row, text="/", bg=row_bg, fg=COLOR_TEXT_MUTED, font=self.fonts.input).place(
            x=x, y=control_y, width=10, height=control_h); x += 14
        self.s_year = SuraEntry(row, width=70, height=control_h, font=self.fonts.input)
        self.s_year.place(x=x, y=control_y); x += 70 + 32

        tk.Label(
            row, text="Fecha fin", bg=row_bg, fg=COLOR_TEXT, font=self.fonts.label, anchor="w",
        ).place(x=x, y=0, width=90, height=row_h)
        x += 90

        self.e_day = SuraEntry(row, width=44, height=control_h, font=self.fonts.input)
        self.e_day.place(x=x, y=control_y); x += 44 + 4
        tk.Label(row, text="/", bg=row_bg, fg=COLOR_TEXT_MUTED, font=self.fonts.input).place(
            x=x, y=control_y, width=10, height=control_h); x += 14
        self.e_month = SuraEntry(row, width=44, height=control_h, font=self.fonts.input)
        self.e_month.place(x=x, y=control_y); x += 44 + 4
        tk.Label(row, text="/", bg=row_bg, fg=COLOR_TEXT_MUTED, font=self.fonts.input).place(
            x=x, y=control_y, width=10, height=control_h); x += 14
        self.e_year = SuraEntry(row, width=70, height=control_h, font=self.fonts.input)
        self.e_year.place(x=x, y=control_y); x += 70 + 24

        tk.Label(
            row, text="Formato DD/MM/AAAA — por defecto, bimestre anterior al actual",
            bg=row_bg, fg=COLOR_TEXT_MUTED, font=self.fonts.status, anchor="w",
        ).place(x=x, y=0, width=row_width - x - 20, height=row_h)

        vcmd_day = (self.root.register(self._validate_digits_len), "%P", "2")
        vcmd_year = (self.root.register(self._validate_digits_len), "%P", "4")
        for e in (self.s_day, self.s_month, self.e_day, self.e_month):
            e.configure_validation(vcmd_day)
        for e in (self.s_year, self.e_year):
            e.configure_validation(vcmd_year)

    def _build_import_row(self, parent: tk.Misc, row_width: int, row_h: int, y: int, row_bg: str) -> None:
        row = tk.Frame(parent, bg=row_bg, bd=0, highlightthickness=0)
        row.place(x=0, y=y, width=row_width, height=row_h)

        tk.Label(
            row, text="Importar", bg=row_bg, fg=COLOR_TEXT, font=self.fonts.label, anchor="w",
        ).place(x=24, y=0, width=110, height=row_h)

        self.var_nat = tk.BooleanVar(value=True)
        self.var_jur = tk.BooleanVar(value=True)
        self.var_gen = tk.BooleanVar(value=True)

        checks = [
            (self.var_nat, "Personas Naturales", 134, 190),
            (self.var_jur, "Personas Jurídicas", 190, 190),
            (self.var_gen, "Tasas sin distinguir persona", 190, 260),
        ]
        x = 134
        for var, text, _w, cw in checks:
            self._checkbox(row, text, var, x, row_h, row_bg)
            x += cw

    def _checkbox(self, parent: tk.Misc, text: str, var: tk.BooleanVar, x: int,
                  row_h: int, row_bg: str) -> tk.Checkbutton:
        cb = tk.Checkbutton(
            parent, text=text, variable=var, bg=row_bg, fg=COLOR_TEXT,
            activebackground=row_bg, activeforeground=COLOR_TEXT,
            selectcolor=COLOR_WHITE, font=self.fonts.checkbox,
            bd=0, highlightthickness=0, cursor="hand2", anchor="w",
        )
        cb.place(x=x, y=0, height=row_h)
        return cb

    def _build_advanced_toggle(self) -> None:
        self.adv_shown = False
        self.adv_toggle_btn = SuraButton(
            self.root, text="Opciones avanzadas ▸", command=self.toggle_advanced,
            variant="secondary", width=200, height=30, font=self.fonts.button,
        )
        self._adv_toggle_y = self._form_end_y + 14
        self.adv_toggle_btn.place(x=MARGIN_X, y=self._adv_toggle_y)

        panel_width = APP_WIDTH - 2 * MARGIN_X
        self._adv_panel_height = 52
        self.adv_panel = tk.Frame(self.root, bg=COLOR_BORDER, bd=0, highlightthickness=0)
        adv_inner = tk.Frame(self.adv_panel, bg=COLOR_WHITE, bd=0, highlightthickness=0)
        adv_inner.place(x=1, y=1, width=panel_width - 2, height=self._adv_panel_height - 2)

        self.var_skip_weekends = tk.BooleanVar(value=True)
        self.var_remove_html = tk.BooleanVar(value=True)
        self.var_show_adv_console = tk.BooleanVar(value=False)
        self.var_detail_log = tk.BooleanVar(value=False)

        row_h = self._adv_panel_height - 2
        x = 24
        self._checkbox(adv_inner, "Omitir sábados y domingos", self.var_skip_weekends, x, row_h, COLOR_WHITE)
        x += 230
        self._checkbox(adv_inner, "Eliminar HTML temporales", self.var_remove_html, x, row_h, COLOR_WHITE)
        x += 220
        cb_console = self._checkbox(adv_inner, "Mostrar consola avanzada", self.var_show_adv_console,
                                     x, row_h, COLOR_WHITE)
        cb_console.configure(command=self.toggle_advanced_console)
        x += 220
        self._checkbox(
            adv_inner, "Log detallado por día (feriados / datos faltantes)",
            self.var_detail_log, x, row_h, COLOR_WHITE,
        )

        self._adv_panel_y = self._adv_toggle_y + 40

    def toggle_advanced(self) -> None:
        self.adv_shown = not self.adv_shown
        if self.adv_shown:
            self.adv_toggle_btn.set_text("Opciones avanzadas ▾")
            self.adv_panel.place(
                x=MARGIN_X, y=self._adv_panel_y, width=APP_WIDTH - 2 * MARGIN_X,
                height=self._adv_panel_height,
            )
        else:
            self.adv_toggle_btn.set_text("Opciones avanzadas ▸")
            self.adv_panel.place_forget()
        self._reposition_log_area()

    def _build_log_area(self) -> None:
        self.log_label = tk.Label(
            self.root, text="Registro de proceso", bg=COLOR_WHITE, fg=COLOR_BLACK,
            font=self.fonts.section, anchor="w",
        )
        self.log_box = SuraLogBox(
            self.root, fonts=self.fonts, width=APP_WIDTH - 2 * MARGIN_X, height=100,
        )

    def _reposition_log_area(self) -> None:
        # El log siempre termina justo antes de la barra de progreso (posición
        # fija); su punto de inicio depende de si el panel de opciones
        # avanzadas está expandido o no, ya que todo el layout usa .place().
        if self.adv_shown:
            top_y = self._adv_panel_y + self._adv_panel_height + 14
        else:
            top_y = self._adv_toggle_y + 40

        label_y = top_y
        log_y = label_y + 26
        height = max(60, LOG_BOTTOM_Y - log_y)

        self.log_label.place(x=MARGIN_X, y=label_y, width=300, height=22)
        self.log_box.reposition(MARGIN_X, log_y, height)

    def _build_action_bar(self) -> None:
        panel_width: int = APP_WIDTH - 2 * MARGIN_X

        self.progress = SuraProgressBar(self.root, width=panel_width, height=9)
        self.progress.place(x=MARGIN_X, y=PROGRESS_Y)

        self.status = tk.Label(
            self.root, text="Listo para iniciar.", bg=COLOR_WHITE, fg=COLOR_TEXT_MUTED,
            font=self.fonts.status, anchor="w",
        )
        self.status.place(x=MARGIN_X, y=BUTTONS_Y + 7, width=560, height=22)

        total_button_width = 150 + 116 + 116 + 10 + 10
        start_x = APP_WIDTH - MARGIN_X - total_button_width

        self.run_btn = SuraButton(
            self.root, text="Ejecutar", command=self.on_run, variant="primary",
            width=150, height=36, font=self.fonts.button,
        )
        self.run_btn.place(x=start_x, y=BUTTONS_Y)

        self.stop_btn = SuraButton(
            self.root, text="Detener", command=self.on_stop, variant="secondary",
            width=116, height=36, font=self.fonts.button,
        )
        self.stop_btn.place(x=start_x + 160, y=BUTTONS_Y)
        self.stop_btn.set_enabled(False)

        self.close_btn = SuraButton(
            self.root, text="Cerrar", command=self.on_close, variant="secondary",
            width=116, height=36, font=self.fonts.button,
        )
        self.close_btn.place(x=start_x + 286, y=BUTTONS_Y)

    # ------------------------------------------------------------- helpers

    def _validate_digits_len(self, proposed: str, maxlen: str) -> bool:
        if proposed == "":
            return True
        if not proposed.isdigit():
            return False
        return len(proposed) <= int(maxlen)

    def _set_default_dates(self) -> None:
        sdate, edate = bimestre_anterior(datetime.now().date())
        self.s_day.set(f"{sdate.day:02d}")
        self.s_month.set(f"{sdate.month:02d}")
        self.s_year.set(f"{sdate.year:04d}")
        self.e_day.set(f"{edate.day:02d}")
        self.e_month.set(f"{edate.month:02d}")
        self.e_year.set(f"{edate.year:04d}")

    def _choose_folder(self) -> None:
        d = filedialog.askdirectory(initialdir=self.folder_entry.get() or str(default_output_dir()))
        if d:
            self.folder_entry.set(d)

    def toggle_advanced_console(self) -> None:
        self._console_mode = "advanced" if self.var_show_adv_console.get() else "simple"
        lines = self._log_advanced_lines if self._console_mode == "advanced" else self._log_simple_lines
        self.log_box.set_content(lines)

    def log_simple(self, msg: str) -> None:
        self._log_simple_lines.append(msg)
        self.status.configure(text=msg)
        if self._console_mode == "simple":
            self.log_box.write(msg)

    def log_advanced(self, msg: str) -> None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {msg}"
        self._log_advanced_lines.append(line)
        if self._console_mode == "advanced":
            self.log_box.write(line)

    # ------------------------------------------------------------- acciones

    def on_run(self) -> None:
        try:
            s_day  = self.s_day.get().strip().zfill(2)
            s_mon  = self.s_month.get().strip().zfill(2)
            s_year = self.s_year.get().strip().zfill(4)
            e_day  = self.e_day.get().strip().zfill(2)
            e_mon  = self.e_month.get().strip().zfill(2)
            e_year = self.e_year.get().strip().zfill(4)
            start  = f"{s_day}/{s_mon}/{s_year}"
            end    = f"{e_day}/{e_mon}/{e_year}"
            parse_ddmmyyyy(start)
            parse_ddmmyyyy(end)
            if parse_ddmmyyyy(end) < parse_ddmmyyyy(start):
                raise ValueError("end_date < start_date")
        except Exception:
            messagebox.showerror(
                "Fecha inválida",
                "Completar las fechas con valores válidos y verificar "
                "que la fecha fin no sea anterior a la fecha inicio.",
            )
            return

        out_base_dir = self.folder_entry.get().strip() or str(default_output_dir())
        self.cancel_event.clear()
        self._worker_status = None
        self.run_btn.set_enabled(False)
        self.stop_btn.set_enabled(True)
        self.progress.set_progress(0.0)
        self.log_simple(f"Iniciando proceso desde {start} hasta {end}.")
        self.log_advanced(f"[INFO] Iniciando proceso desde {start} hasta {end}.")

        generate_detail_log = self.var_detail_log.get()

        self.worker = threading.Thread(
            target=self._worker, args=(start, end, out_base_dir, generate_detail_log), daemon=True,
        )
        self.worker.start()
        self.root.after(300, self._poll_worker)

    def _worker(self, start: str, end: str, out_base_dir: str, generate_detail_log: bool) -> None:
        status = "error_excepcion"
        try:
            status = run_date_range(
                start, end,
                out_base_dir        = out_base_dir,
                log_fn               = lambda s: self.root.after(0, lambda: self.log_advanced(s)),
                simple_log_fn        = lambda s: self.root.after(0, lambda: self.log_simple(s)),
                import_natural       = self.var_nat.get(),
                import_juridica      = self.var_jur.get(),
                import_general       = self.var_gen.get(),
                remove_html_after    = self.var_remove_html.get(),
                skip_weekends        = self.var_skip_weekends.get(),
                cancel_event         = self.cancel_event,
                generate_detail_log  = generate_detail_log,
                progress_fn          = lambda f: self.root.after(0, lambda: self.progress.set_progress(f)),
            )
        except Exception as e:
            self.root.after(0, lambda: self.log_advanced(f"[ERROR] Excepción no controlada en el proceso: {e}"))
            self.root.after(0, lambda: self.log_advanced(traceback.format_exc()))
            self.root.after(0, lambda: self.log_simple("Se produjo un error inesperado. Revisar la consola avanzada."))
        self._worker_status = status

    def _poll_worker(self) -> None:
        if self.worker and self.worker.is_alive():
            self.root.after(500, self._poll_worker)
            return

        self.run_btn.set_enabled(True)
        self.stop_btn.set_enabled(False)

        mensajes = {
            "completado":       "Proceso terminado.",
            "cancelado":        "Proceso detenido por el usuario.",
            "error_chrome":     "Proceso finalizado con error: no se pudo iniciar Chrome.",
            "error_navegacion": "Proceso finalizado con error: no se pudo cargar la página de la SBS.",
            "error_excepcion":  "Proceso finalizado con un error inesperado. Revisar la consola avanzada.",
        }
        msg = mensajes.get(self._worker_status, "Proceso terminado.")
        self.log_simple(msg)
        self.log_advanced(f"[INFO] {msg}")

    def on_stop(self) -> None:
        # No reactivar "Ejecutar" aquí: el hilo (y Chrome) siguen vivos hasta
        # que terminen la fecha en curso y salgan del bucle; _poll_worker se
        # encarga de reactivar los botones recién cuando el hilo finalice.
        self.cancel_event.set()
        self.stop_btn.set_enabled(False)
        self.log_simple("Deteniendo el proceso — esperando a que finalice la fecha en curso.")
        self.log_advanced("[INFO] Solicitud de detener recibida por el usuario.")

    def on_close(self) -> None:
        if self.worker and self.worker.is_alive():
            if not messagebox.askyesno(
                "Proceso en curso", "Hay un proceso en ejecución. ¿Detener el proceso y salir?",
            ):
                return
            self.cancel_event.set()
            self.root.after(200, self._wait_and_destroy)
        else:
            self.root.destroy()

    def _wait_and_destroy(self) -> None:
        if self.worker and self.worker.is_alive():
            self.root.after(200, self._wait_and_destroy)
        else:
            self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()

# ==================== ENTRYPOINT ====================


def main() -> None:
    app = TasasApp()
    app.run()


if __name__ == "__main__":
    main()